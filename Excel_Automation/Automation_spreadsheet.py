'''Project 1: Automation: program that process the thoussands of spredsheets in under a sec
if we go manually, it takes weeks and months but python program it can takes seconds.'''



'''
transaction_id	product_id	price 
1001	        1	        $5.95
1002	        2	        $6.95
1003	        3	        $7.95

'''

'''Project:'''
import openpyxl as xl
wb = xl.load_workbook("transactions.xlsx")
from openpyxl.chart import BarChart, Reference
#
#
sheet = wb['Sheet1']
#
# cell = sheet['a1']
# or
cell = sheet.cell(2, 3)  # braces represent-->(row, Column)
print(cell.value)
#
print(sheet.max_row)
# print(sheet.max_column)


for row in range(2, sheet.max_row +1):  # 2 3 4
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

sheet.cell(row=1, column=4, value='corrected_price')

for row in sheet.values:
    print(row)

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart =BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'b7')

wb.save('transactions1.xlsx')

Reference()